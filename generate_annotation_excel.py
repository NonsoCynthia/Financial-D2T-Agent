"""
generate_annotation_excel.py

Builds a human-evaluation Excel workbook for the financial NLG annotation study.

Reads result files directly from the four NLG output directories:

  BR conditions (all 24 reports each)  →  48 PT-BR reports total
  EN conditions (all 14 reports each)  →  28 EN reports total
  Grand total: 76 reports, shuffled and blinded as R-01 … R-76

Workbook structure:
  Guide      — condensed 1-7 criteria reference (keep open while scoring)
  Scores     — rating sheet, hyperlinked R-XX IDs, 6 score columns + Comments
  R-01…R-76  — one sheet per report: INPUT DATA (col A) | GENERATED REPORT (col B)
  _Key       — hidden mapping R-XX → condition/category/file (not for raters)

Input format per language:
  EN  — SPO triples from the JSON `data` field, formatted as Subject | Predicate | Value
  PT-BR — structured per-ticker query text from the JSON `query` field

Usage:
  /home/chinonso/anaconda3/bin/python3 generate_annotation_excel.py
"""

from __future__ import annotations

import hashlib
import json
import random
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths and sampling configuration
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).parent
RESULTS_ROOT = PROJECT_ROOT / "results"
HUMAN_EVAL_DIR = RESULTS_ROOT / "validation" / "human_eval_multi_model_robustness"
OUTPUT_EXCEL = HUMAN_EVAL_DIR / "human_evaluation_annotation.xlsx"

SEED = 20260512

# Each entry: (condition_label, category_path, language, n_samples)
# n_samples=None means take all available files.
CONDITIONS: list[tuple[str, str, str, int | None]] = [
    (
        "BR e2e",
        "nlg_brazilian_manager/e2e",
        "pt-BR",
        None,   # all 24
    ),
    (
        "BR default",
        "nlg_brazilian_manager/default",
        "pt-BR",
        None,   # all 24
    ),
    (
        "US e2e (reflection)",
        "nlg/final_report2025_us/gpt-5/workflow_True/openai/gpt-5/en/e2e",
        "en",
        None,   # all 14
    ),
    (
        "US default (reflection)",
        "nlg/final_report2025_us/gpt-5/workflow_True/openai/gpt-5/en/default",
        "en",
        None,   # all 14
    ),
]

# Pilot study: 2 samples each from no-reflection e2e and no-reflection pipeline (default)
PILOT_CONDITIONS: list[tuple[str, str, str, int]] = [
    (
        "US e2e (no reflection)",
        "nlg/final_report2025_us/gpt-5/workflow_False/openai/gpt-5/en/e2e",
        "en",
        2,
    ),
    (
        "US default (no reflection)",
        "nlg/final_report2025_us/gpt-5/workflow_False/openai/gpt-5/en/default",
        "en",
        2,
    ),
]

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

EN_BG        = "DDEEFF"   # light blue — English
PT_BG        = "E6F4EA"   # light green — PT-BR
HEADER_DARK  = "2F5496"   # dark blue — title bars
HEADER_MID   = "4472C4"   # medium blue — EN headers
HEADER_PT    = "2D6A4F"   # dark green — PT-BR headers
SCORE_FILL   = "FFF9C4"   # pale yellow — empty score cells
PILOT_HEADER = "A04000"   # burnt orange — pilot study sheets
PILOT_BG     = "FFF0E6"   # pale orange — pilot score rows


def thin_border(color: str = "BBBBBB") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def solid(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def cfont(bold: bool = False, size: int = 10, color: str = "000000",
          italic: bool = False, name: str = "Calibri") -> Font:
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)


def wrap_top(h: str = "left") -> Alignment:
    return Alignment(wrap_text=True, vertical="top", horizontal=h)


def center_align() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def stable_item_id(condition: str, filename: str) -> str:
    digest = hashlib.sha1(f"{condition}|{filename}".encode()).hexdigest()[:10]
    return f"HE-{digest}"


def load_report(json_path: Path, condition: str, language: str) -> dict[str, Any]:
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    txt_path = json_path.with_suffix(".txt")
    generated_text = (
        payload.get("generated_text")
        or payload.get("final_response")
        or (txt_path.read_text(encoding="utf-8") if txt_path.exists() else "")
    ).strip()

    meta = payload.get("sample_metadata") or {}
    analysis_date = str(
        payload.get("analysis_date") or meta.get("analysis_date") or "unknown"
    )
    return {
        "human_item_id": stable_item_id(condition, json_path.name),
        "condition": condition,
        "language": language,
        "analysis_date": analysis_date,
        "report_name": json_path.stem,
        "json_path": json_path,
        "generated_text": generated_text,
    }


def format_us_input(json_path: Path) -> str:
    """Format SPO triples as Subject | Predicate | Value, one per line.

    Handles two JSON schemas:
      e2e format    — triples in top-level `data`
      default format — triples in top-level `data_input` or `sample_metadata.data_input`
    """
    try:
        payload = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return f"(Failed to load JSON: {exc})"

    meta = payload.get("sample_metadata") or {}
    analysis_date = str(payload.get("analysis_date") or meta.get("analysis_date") or "")
    report_subject = f"M_SMRG_{analysis_date}" if analysis_date else "M_SMRG"
    header_triples = [
        (report_subject, "report_type", "Monthly Equity Review"),
        (report_subject, "analysis_date", analysis_date),
        (report_subject, "price_reference_date", analysis_date),
        (report_subject, "coverage_window_end_date", str(payload.get("end_date") or "")),
        (report_subject, "investment_horizon_months", str(payload.get("horizon_months") or "")),
        (
            report_subject,
            "analyst_note",
            "This report is generated from structured financial data. All recommendations are model-derived. This document does not constitute regulated investment advice. Past performance is not indicative of future results.",
        ),
    ]
    triples = (
        payload.get("data")
        or payload.get("data_input")
        or meta.get("data_input")
        or []
    )
    lines = [f"{s} | {p} | {v}" for s, p, v in header_triples if v]
    lines.extend(f"{s} | {p} | {v}" for s, p, v in triples if len((s, p, v)) == 3)
    return "\n".join(lines) or "(no triples found)"


def format_br_input(json_path: Path) -> str:
    """Return the structured per-ticker query text from the BR manager JSON.

    Handles two JSON schemas:
      e2e format     — text in top-level `query`
      default format — text in `sample_metadata.prompt_context`
    """
    try:
        payload = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return f"(Failed to load JSON: {exc})"
    meta = payload.get("sample_metadata") or {}
    query = (
        payload.get("query")
        or payload.get("input")
        or payload.get("prompt")
        or meta.get("prompt_context")
        or ""
    )
    return str(query).strip()


def collect_all_reports() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Return (en_records, pt_records) each sorted by analysis date."""
    en_records: list[dict[str, Any]] = []
    pt_records: list[dict[str, Any]] = []

    for condition, category, language, n_samples in CONDITIONS:
        folder = RESULTS_ROOT / category
        if not folder.exists():
            print(f"  WARNING: folder not found — {folder}")
            continue

        json_files = sorted(
            p for p in folder.glob("*.json")
            if "sequence_summary" not in p.name
        )

        if n_samples is not None:
            rng_local = random.Random(SEED)
            rng_local.shuffle(json_files)
            json_files = json_files[:n_samples]

        records = [load_report(p, condition, language) for p in json_files]
        if language == "en":
            en_records.extend(records)
        else:
            pt_records.extend(records)
        print(f"  {condition}: {len(records)} reports  ({language})")

    # Within each language group, shuffle for blinded order
    rng = random.Random(SEED)
    rng.shuffle(en_records)
    rng.shuffle(pt_records)
    return en_records, pt_records


def collect_pilot_reports() -> list[dict[str, Any]]:
    """Return exactly n_samples records per pilot condition, seeded for reproducibility."""
    pilot_records: list[dict[str, Any]] = []
    for condition, category, language, n_samples in PILOT_CONDITIONS:
        folder = RESULTS_ROOT / category
        if not folder.exists():
            print(f"  WARNING (pilot): folder not found — {folder}")
            continue
        json_files = sorted(
            p for p in folder.glob("*.json")
            if "sequence_summary" not in p.name
        )
        rng_local = random.Random(SEED)
        rng_local.shuffle(json_files)
        json_files = json_files[:n_samples]
        records = [load_report(p, condition, language) for p in json_files]
        pilot_records.extend(records)
        print(f"  [pilot] {condition}: {len(records)} reports  ({language})")
    return pilot_records

# ---------------------------------------------------------------------------
# Guide sheet
# ---------------------------------------------------------------------------

CRITERIA = [
    (
        "No Omissions",
        "Whether the report includes ALL information present in the input data "
        "(financial indicators, recommendations, target prices, balance sheet figures, qualitative facts).\n\n"
        "5 — All facts from the input are present and accurately represented. Nothing left out.\n"
        "4 — Nearly all facts covered. Only minor or peripheral items are missing.\n"
        "3 — Most facts present, but noticeable gaps exist across one or more tickers.\n"
        "2 — Several important facts missing. Coverage substantially incomplete.\n"
        "1 — The report fails to cover the vast majority of input facts. Severely deficient.",
    ),
    (
        "No Additions",
        "Whether the report contains ONLY information that came from the input data. "
        "Any figure, claim, prediction, or assertion not in the input counts as an addition.\n\n"
        "5 — Every claim and figure fully supported by the input. No additions detected.\n"
        "4 — Nearly all content traceable to input. Only very minor rhetorical additions.\n"
        "3 — A few additions present, but majority of claims are grounded in the input.\n"
        "2 — Several additions present. Substantial portion goes beyond the input.\n"
        "1 — Numerous invented or unsupported claims. Hallucinations are pervasive.",
    ),
    (
        "Fluency",
        "How naturally and smoothly the report reads: sentence flow, prose rhythm, "
        "whether it sounds like a competent human author. Technical financial vocabulary is expected and should not be penalised.\n\n"
        "5 — Exceptional fluency throughout. Prose is polished and completely natural.\n"
        "4 — Reads naturally in most places. Minor issues do not impede reading.\n"
        "3 — Generally readable, but some phrases feel stilted, repetitive, or mechanical.\n"
        "2 — Frequently unnatural. Awkward phrasing regularly impedes reading.\n"
        "1 — Very unnatural throughout. Awkward phrasing makes text difficult to read.",
    ),
    (
        "Grammaticality",
        "Whether the report is grammatically correct: sentence structure, subject-verb agreement, "
        "punctuation, tense consistency, article use, typographical errors. "
        "For PT-BR reports, standard written Brazilian Portuguese grammar applies. "
        "Score separately from Fluency.\n\n"
        "5 — Grammatically flawless throughout. No errors of any kind.\n"
        "4 — Minor, infrequent errors that do not impede comprehension.\n"
        "3 — Several errors present, but overall meaning is clear.\n"
        "2 — Frequent errors regularly interrupt comprehension.\n"
        "1 — Pervasive grammatical errors. Text is difficult to parse.",
    ),
    (
        "Coherence",
        "Whether the report is well-organised and logically connected: ordering of information, "
        "smooth transitions between sections and tickers, internal consistency, reads as a unified whole.\n\n"
        "5 — Exceptionally well-organised. All sections flow logically as a cohesive whole.\n"
        "4 — Well-organised in most places. Minor structural or transitional issues do not impede reading.\n"
        "3 — Generally coherent, but some sections feel disconnected or ordering could be improved.\n"
        "2 — Poorly organised. Frequently feels fragmented or internally inconsistent.\n"
        "1 — Incoherent. Lacks logical structure; reads as disconnected statements.",
    ),
    (
        "Analytical Utility",
        "Whether the report goes beyond mechanically listing input facts to offer genuinely useful analysis: "
        "synthesis across tickers, meaningful comparisons, identification of patterns or risks, coherent investment framing. "
        "Score on analytical quality only — not on financial accuracy.\n\n"
        "5 — Exceptionally analytical. Synthesis is comprehensive, nuanced, and highly actionable.\n"
        "4 — Clearly analytical. Meaningful comparisons and interpretations appear consistently.\n"
        "3 — Some analytical content present, but substantial portion remains descriptive.\n"
        "2 — Almost no analysis. Nearly all content is a direct restatement of input facts.\n"
        "1 — Pure recitation of input facts. No meaningful analysis attempted.",
    ),
]

REMINDERS = [
    "Score each criterion independently. Your overall impression should not inflate or deflate individual scores.",
    "No Omissions and No Additions are assessed relative to the input data only — not against real-world financial knowledge.",
    "Fluency and Grammaticality are intrinsic properties of the text itself. That is only the text is used here to determine the text quality.",
    "Coherence is assessed on logical structure and internal consistency, not whether investment stances are financially sound.",
    "Analytical Utility rewards synthesis and interpretation, producing new insights which are potentialy actionable. However it does not imply not real-world financial correctness.",
    "If missing data are explicitly acknowledged in the input or report, do not penalise the report for not inventing those values.",
    "Evaluators should only assess reports in languages in which they are proficient.",
]


def build_guide_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Guide")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95
    ws.sheet_tab_color = "375623"

    row = 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1,
                value="Human Evaluation Guide — Financial NLG Study")
    c.font = cfont(bold=True, size=15, color="FFFFFF")
    c.fill = solid(HEADER_DARK)
    c.alignment = center_align()
    ws.row_dimensions[row].height = 32
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(
        row=row, column=1,
        value="Scale: 1 = lowest  ·  5 = highest  ·  Whole numbers only  "
              "·  When torn between two adjacent scores, choose the lower one",
    )
    c.font = cfont(bold=True, italic=True, size=10, color="FFFFFF")
    c.fill = solid(HEADER_MID)
    c.alignment = center_align()
    ws.row_dimensions[row].height = 18
    row += 1

    row += 1  # blank

    for name, description in CRITERIA:
        a = ws.cell(row=row, column=1, value=name)
        a.font = cfont(bold=True, size=10, color="FFFFFF")
        a.fill = solid(HEADER_MID)
        a.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        b = ws.cell(row=row, column=2, value=description)
        b.font = Font(name="Calibri", size=9)
        b.fill = solid("F4F8FF")
        b.alignment = wrap_top()

        ws.row_dimensions[row].height = 120
        row += 1

    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1, value="Key Reminders")
    c.font = cfont(bold=True, size=11, color="FFFFFF")
    c.fill = solid(HEADER_DARK)
    c.alignment = center_align()
    ws.row_dimensions[row].height = 20
    row += 1

    for reminder in REMINDERS:
        ws.cell(row=row, column=1, value="•").font = cfont(bold=True, size=11)
        ws.cell(row=row, column=1).alignment = center_align()
        ws.cell(row=row, column=2, value=reminder).font = Font(name="Calibri", size=10)
        ws.cell(row=row, column=2).alignment = wrap_top()
        ws.row_dimensions[row].height = 18
        row += 1

# ---------------------------------------------------------------------------
# Scores sheet
# ---------------------------------------------------------------------------

SCORE_HEADERS = [
    "Report ID",
    "Language",
    "Condition",
    "No Omissions\n(1–5)",
    "No Additions\n(1–5)",
    "Fluency\n(1–5)",
    "Grammaticality\n(1–5)",
    "Coherence\n(1–5)",
    "Analytical Utility\n(1–5)",
    "Comments",
]

COL_WIDTHS_SCORES = {1: 11, 2: 11, 3: 26, 4: 16, 5: 16, 6: 12, 7: 16, 8: 13, 9: 18, 10: 45}


def build_scores_sheet(
    wb: openpyxl.Workbook,
    records: list[dict[str, Any]],
    report_ids: list[str],
) -> None:
    ws = wb.create_sheet("Scores")
    ws.freeze_panes = "D2"
    ws.sheet_tab_color = HEADER_DARK

    for col, w in COL_WIDTHS_SCORES.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    for col, h in enumerate(SCORE_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = cfont(bold=True, size=10, color="FFFFFF")
        c.fill = solid(HEADER_DARK)
        c.alignment = center_align()
        c.border = thin_border()
    ws.row_dimensions[1].height = 38

    link_font = Font(name="Calibri", size=10, bold=True,
                     color="0563C1", underline="single")

    for i, record in enumerate(records):
        excel_row = i + 2
        rid = report_ids[i]
        lang = record["language"]
        lang_label = "EN" if lang == "en" else "PT-BR"
        row_bg = solid(EN_BG if lang == "en" else PT_BG)

        id_cell = ws.cell(row=excel_row, column=1, value=rid)
        id_cell.font = link_font
        id_cell.hyperlink = f"#{rid}!A1"
        id_cell.alignment = center_align()
        id_cell.fill = row_bg
        id_cell.border = thin_border()

        lang_cell = ws.cell(row=excel_row, column=2, value=lang_label)
        lang_cell.font = cfont(bold=True, size=10)
        lang_cell.alignment = center_align()
        lang_cell.fill = row_bg
        lang_cell.border = thin_border()

        cond_cell = ws.cell(row=excel_row, column=3, value=record["condition"])
        cond_cell.font = cfont(size=10)
        cond_cell.alignment = center_align()
        cond_cell.fill = row_bg
        cond_cell.border = thin_border()

        for col in range(4, 10):
            c = ws.cell(row=excel_row, column=col, value=None)
            c.fill = solid(SCORE_FILL)
            c.alignment = center_align()
            c.border = thin_border()
            c.font = cfont(size=11)

        ws.cell(row=excel_row, column=10).alignment = wrap_top()
        ws.cell(row=excel_row, column=10).border = thin_border()
        ws.row_dimensions[excel_row].height = 20

# ---------------------------------------------------------------------------
# Report sheets
# ---------------------------------------------------------------------------

def build_report_sheet(
    wb: openpyxl.Workbook,
    record: dict[str, Any],
    report_id: str,
    condition: str | None = None,
) -> None:
    ws = wb.create_sheet(report_id)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 95

    lang = record["language"]
    is_en = lang == "en"
    lang_label = "English" if is_en else "Brazilian Portuguese"
    is_pilot = condition is not None
    header_color = PILOT_HEADER if is_pilot else (HEADER_MID if is_en else HEADER_PT)
    ws.sheet_tab_color = header_color

    json_path: Path = record["json_path"]

    if is_en:
        input_text = format_us_input(json_path)
    else:
        input_text = format_br_input(json_path)

    output_text = record["generated_text"]

    # Row 1: banner
    ws.merge_cells("A1:B1")
    banner_text = f"{report_id}   —   {lang_label}"
    title = ws.cell(row=1, column=1, value=banner_text)
    title.font = cfont(bold=True, size=14, color="FFFFFF")
    title.fill = solid(header_color)
    title.alignment = center_align()
    ws.row_dimensions[1].height = 30

    # Row 2: back link
    back_sheet = "Pilot_Scores" if is_pilot else "Scores"
    back = ws.cell(row=2, column=1, value=f"← Back to {back_sheet}")
    back.font = Font(name="Calibri", size=9, color="0563C1", underline="single")
    back.hyperlink = f"#{back_sheet}!A1"
    back.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    # Row 3: blank
    ws.row_dimensions[3].height = 5

    # Row 4: column headers
    for col, label in ((1, "INPUT DATA"), (2, "GENERATED REPORT")):
        h = ws.cell(row=4, column=col, value=label)
        h.font = cfont(bold=True, size=11, color="FFFFFF")
        h.fill = solid(header_color)
        h.alignment = center_align()
    ws.row_dimensions[4].height = 22

    # Row 5: content
    in_cell = ws.cell(row=5, column=1, value=input_text)
    in_cell.font = Font(name="Courier New", size=8.5)
    in_cell.alignment = wrap_top()
    in_cell.fill = solid("FAFAFA")

    out_cell = ws.cell(row=5, column=2, value=output_text)
    out_cell.font = Font(name="Calibri", size=10)
    out_cell.alignment = wrap_top()
    out_cell.fill = solid("FFFFFF")

    input_lines = input_text.count("\n") + 1
    output_lines = max(output_text.count("\n") + 1, len(output_text) // 115)
    ws.row_dimensions[5].height = min(max(input_lines, output_lines) * 13, 409)

# ---------------------------------------------------------------------------
# Private key sheet
# ---------------------------------------------------------------------------

KEY_WARNING = "RESEARCHER USE ONLY — DO NOT SHARE THIS SHEET WITH RATERS"
KEY_WARNING_FILL = "C00000"   # dark red


def build_key_sheet(
    wb: openpyxl.Workbook,
    records: list[dict[str, Any]],
    report_ids: list[str],
) -> None:
    ws = wb.create_sheet("_Key")
    ws.sheet_tab_color = KEY_WARNING_FILL

    key_cols = ["report_id", "human_item_id", "condition", "language",
                "analysis_date", "report_name", "json_path"]

    # Row 1: prominent warning
    ws.merge_cells(f"A1:{get_column_letter(len(key_cols))}1")
    warn = ws.cell(row=1, column=1, value=KEY_WARNING)
    warn.font = cfont(bold=True, size=12, color="FFFFFF")
    warn.fill = solid(KEY_WARNING_FILL)
    warn.alignment = center_align()
    ws.row_dimensions[1].height = 24

    # Row 2: column headers
    for col, h in enumerate(key_cols, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = cfont(bold=True, color="FFFFFF")
        c.fill = solid(HEADER_DARK)
        ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 2, 16)

    for i, record in enumerate(records):
        row = i + 3
        values = [
            report_ids[i],
            record["human_item_id"],
            record["condition"],
            record["language"],
            record["analysis_date"],
            record["report_name"],
            str(record["json_path"].relative_to(PROJECT_ROOT)),
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val)

# ---------------------------------------------------------------------------
# Pilot scores and key sheets
# ---------------------------------------------------------------------------

PILOT_SCORE_HEADERS = [
    "Report ID",
    "Language",
    "Condition",
    "No Omissions\n(1–5)",
    "No Additions\n(1–5)",
    "Fluency\n(1–5)",
    "Grammaticality\n(1–5)",
    "Coherence\n(1–5)",
    "Analytical Utility\n(1–5)",
    "Comments",
]
# col index → width  (Condition inserted at col 3, scores shift to 4–9, Comments = 10)
PILOT_COL_WIDTHS = {1: 11, 2: 11, 3: 26, 4: 16, 5: 16, 6: 12, 7: 16, 8: 13, 9: 18, 10: 45}


def build_pilot_scores_sheet(
    wb: openpyxl.Workbook,
    pilot_records: list[dict[str, Any]],
    pilot_ids: list[str],
) -> None:
    ws = wb.create_sheet("Pilot_Scores")
    ws.freeze_panes = "D2"
    ws.sheet_tab_color = PILOT_HEADER

    for col, w in PILOT_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Banner explaining the pilot purpose
    ws.merge_cells(f"A1:{get_column_letter(len(PILOT_SCORE_HEADERS))}1")
    banner = ws.cell(
        row=1, column=1,
        value=(
            "PILOT STUDY — Score these 4 items first to calibrate your judgements. "
            "Do not proceed to the main Scores sheet until all 4 are complete."
        ),
    )
    banner.font = cfont(bold=True, size=10, color="FFFFFF")
    banner.fill = solid(PILOT_HEADER)
    banner.alignment = center_align()
    ws.row_dimensions[1].height = 22

    # Column headers on row 2
    for col, h in enumerate(PILOT_SCORE_HEADERS, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = cfont(bold=True, size=10, color="FFFFFF")
        c.fill = solid(PILOT_HEADER)
        c.alignment = center_align()
        c.border = thin_border()
    ws.row_dimensions[2].height = 38

    link_font = Font(name="Calibri", size=10, bold=True,
                     color="0563C1", underline="single")

    for i, record in enumerate(pilot_records):
        excel_row = i + 3
        pid = pilot_ids[i]
        lang_label = "EN" if record["language"] == "en" else "PT-BR"
        condition_label = record["condition"]

        id_cell = ws.cell(row=excel_row, column=1, value=pid)
        id_cell.font = link_font
        id_cell.hyperlink = f"#{pid}!A1"
        id_cell.alignment = center_align()
        id_cell.fill = solid(PILOT_BG)
        id_cell.border = thin_border()

        lang_cell = ws.cell(row=excel_row, column=2, value=lang_label)
        lang_cell.font = cfont(bold=True, size=10)
        lang_cell.alignment = center_align()
        lang_cell.fill = solid(PILOT_BG)
        lang_cell.border = thin_border()

        cond_cell = ws.cell(row=excel_row, column=3, value=condition_label)
        cond_cell.font = cfont(bold=True, size=10)
        cond_cell.alignment = center_align()
        cond_cell.fill = solid(PILOT_BG)
        cond_cell.border = thin_border()

        for col in range(4, 10):
            c = ws.cell(row=excel_row, column=col, value=None)
            c.fill = solid(SCORE_FILL)
            c.alignment = center_align()
            c.border = thin_border()
            c.font = cfont(size=11)

        ws.cell(row=excel_row, column=10).alignment = wrap_top()
        ws.cell(row=excel_row, column=10).border = thin_border()
        ws.row_dimensions[excel_row].height = 20


def build_pilot_key_sheet(
    wb: openpyxl.Workbook,
    pilot_records: list[dict[str, Any]],
    pilot_ids: list[str],
) -> None:
    ws = wb.create_sheet("_Pilot_Key")
    ws.sheet_tab_color = KEY_WARNING_FILL

    key_cols = ["report_id", "human_item_id", "condition", "language",
                "analysis_date", "report_name", "json_path"]

    # Row 1: prominent warning
    ws.merge_cells(f"A1:{get_column_letter(len(key_cols))}1")
    warn = ws.cell(row=1, column=1, value=KEY_WARNING)
    warn.font = cfont(bold=True, size=12, color="FFFFFF")
    warn.fill = solid(KEY_WARNING_FILL)
    warn.alignment = center_align()
    ws.row_dimensions[1].height = 24

    # Row 2: column headers
    for col, h in enumerate(key_cols, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = cfont(bold=True, color="FFFFFF")
        c.fill = solid(PILOT_HEADER)
        ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 2, 16)

    for i, record in enumerate(pilot_records):
        row = i + 3
        values = [
            pilot_ids[i],
            record["human_item_id"],
            record["condition"],
            record["language"],
            record["analysis_date"],
            record["report_name"],
            str(record["json_path"].relative_to(PROJECT_ROOT)),
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("Collecting pilot reports …")
    pilot_records = collect_pilot_reports()
    n_pilot = len(pilot_records)
    pilot_ids = [f"Pilot_{str(i + 1).zfill(2)}" for i in range(n_pilot)]

    print("\nCollecting main study reports …")
    en_records, pt_records = collect_all_reports()
    records = en_records + pt_records
    n = len(records)

    en_count = len(en_records)
    pt_count = len(pt_records)
    print(f"\nTotal: {n} main reports  (EN={en_count}, PT-BR={pt_count})")
    print(f"Pilot: {n_pilot} reports  (Pilot_01–Pilot_{str(n_pilot).zfill(2)})")

    en_ids = [f"E-{str(i + 1).zfill(2)}" for i in range(en_count)]
    pt_ids = [f"BR-{str(i + 1).zfill(2)}" for i in range(pt_count)]
    report_ids = en_ids + pt_ids

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("Building Guide sheet …")
    build_guide_sheet(wb)

    print("Building Pilot_Scores sheet …")
    build_pilot_scores_sheet(wb, pilot_records, pilot_ids)

    print("Building pilot report sheets …")
    for i, record in enumerate(pilot_records):
        build_report_sheet(wb, record, pilot_ids[i], condition=record["condition"])

    print("Building Scores sheet …")
    build_scores_sheet(wb, records, report_ids)

    print("Building main report sheets …")
    for i, record in enumerate(records):
        build_report_sheet(wb, record, report_ids[i])
        if (i + 1) % 10 == 0:
            print(f"  {i + 1}/{n} done")

    print("Building _Key sheet …")
    build_key_sheet(wb, records, report_ids)

    print("Building _Pilot_Key sheet …")
    build_pilot_key_sheet(wb, pilot_records, pilot_ids)

    HUMAN_EVAL_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_EXCEL)

    print(f"\nWorkbook saved: {OUTPUT_EXCEL.relative_to(PROJECT_ROOT)}")
    print(f"  Sheets: Guide | Pilot_Scores | Pilot_01–Pilot_{str(n_pilot).zfill(2)} | Scores | E-01–E-{str(en_count).zfill(2)} | BR-01–BR-{str(pt_count).zfill(2)} | _Key (hidden) | _Pilot_Key (hidden)")
    print(f"  Pilot: {n_pilot} items (orange, pages Pilot_01–Pilot_{str(n_pilot).zfill(2)})")
    print(f"  EN={en_count} (blue, pages E-01–E-{str(en_count).zfill(2)})")
    print(f"  PT-BR={pt_count} (green, pages BR-01–BR-{str(pt_count).zfill(2)})")
    print(f"  Do NOT share the _Key or _Pilot_Key sheets with raters.")


if __name__ == "__main__":
    main()
